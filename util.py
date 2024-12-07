import openpyxl
import calendar
import os
import datetime
import numpy as np
from scipy.stats import norm
from scipy.linalg import cholesky
from scipy.stats import norm


def create_new_xlsx_monthly_dates(load_data, filename, secondTime = 0):

    def month_increment(start_date, num_months):
        # Calculate the new month and year
        new_month = (start_date.month + num_months - 1) % 12 + 1
        new_year = start_date.year + (start_date.month + num_months - 1) // 12
        
        # Calculate the last day of the new month
        last_day_of_month = calendar.monthrange(new_year, new_month)[1]
        
        # Ensure the new day is the last valid day of the new month if the original day doesn't exist in the new month
        new_day = min(start_date.day, last_day_of_month)
        return datetime.date(new_year, new_month, new_day)
    start_date = datetime.date(2024, 4, 30)
    monthly_dates = [month_increment(start_date, i) for i in range(load_data.shape[0])]

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(['Date', '^GSPC', '^ACWX', '^GLAB.L'])

    for i, row in enumerate(load_data):
        ws.append([monthly_dates[i].strftime('%Y-%m-%d')] + row.tolist())
    wb.save(filename)

def binary_to_asset_values_qc(binary, num_assets, expected_returns, cov_matrix):
    # Convert the binary string to a list of integers
    binary_list = [int(bit) for bit in binary]
    
    # Calculate the asset values
    asset_values = []
    for i in range(num_assets):
        asset_value = expected_returns[i]
        for j in range(num_assets):
            asset_value += binary_list[j] * cov_matrix[i, j]
        asset_values.append(asset_value)
    
    return asset_values

def generate_quantum_normal_distribution_all_assets(expected_returns, cov_matrix, num_qubits, stddev):
    # Calculate the number of assets
    num_assets = len(expected_returns)
    
    # Initialize the quantum circuits
    qc_array = []
    for i in range(num_assets):
        qc = generate_quantum_normal_distribution(expected_returns[i], cov_matrix[i, i], num_qubits[i], stddev[i])
        qc_array.append(qc)
    
    return qc_array

def binary_to_asset_values_timeseries(time_series, mu, sigma):
    # Calculate the number of assets
    elements1 = [sublist[0] for sublist in time_series]
    elements2 = [sublist[1] for sublist in time_series]
    elements3 = [sublist[2] for sublist in time_series]

    # Calculate the average and standard deviation of the elements
    average1 = np.mean(elements1)
    average2 = np.mean(elements2)
    average3 = np.mean(elements3)
    stddev1 = np.std(elements1)
    stddev2 = np.std(elements2)
    stddev3 = np.std(elements3)

    # Normalize the elements
    for row in time_series:
        row[0] = ((row[0]- average1) / stddev1 ) * np.sqrt(sigma[0][0]) + mu[0]
        row[1] = ((row[1]- average2) / stddev2 ) * np.sqrt(sigma[1][1]) + mu[1]
        row[2] = ((row[2]- average3) / stddev3 ) * np.sqrt(sigma[2][2]) + mu[2]

    return time_series

